from json import load
from  openpyxl import load_workbook
from utils.check_date_format import check_date_formatISO8601
from utils.custom_exceptions import CustomError

def run(file):
    datas =  {
        'usersInfo':[],
        'useValidation':set()#i used set cus it removes duplicate
    }
    book =  load_workbook(file)
    # so this loop helps me get all the sheet and store the the currentSheet with the sheet we workin on
    # so
    currentSheet = book.sheetnames[0]#we only looking at sheet one in this code so all the second level databae should be in one sheet
    rows = book[currentSheet].rows
    headers = [cell.value for cell in next(rows)]
    # this loop store the headers flagged as --valid
    for head in headers:
        if head is not None:
            if(len(head.split('--'))==2):
                splittedHead=head.split('--')
                if(splittedHead[1].lower()== 'valid'):
                    datas['useValidation'].add(splittedHead[0])



    # cell_line = 0
    # for row in rows:
    #     data =dict()
    #     for title,cell in zip(headers,row):
    #         if title is not None:
    #             if(len(title.split('--'))==2):
    #                 'we dont want to store titleName --valid so her we removeing th --valid that if it exist'
    #                 splittedTitle=title.split('--')
                
    #                 if(splittedTitle[1].lower()== 'valid'):
    #                     validValue =  cell.value
    #                     #the if statement below is to ensure the membership_id is not a number or float that it a strinng 
    #                     # if(type(cell.value) == float):
    #                     #     validValue = str(int(cell.value))
    #                     if(type(cell.value) == str):
    #                         validValue = str(int(cell.value))
                        
    #                     data[splittedTitle[0]] =validValue

    #                 else:
    #                     data[title] = cell.value

                
    #             else:
    #                 data[title] = cell.value


    cell_line = 0
    for row in rows:
        data =dict()
        for title,cell in zip(headers,row):
            if title is not None:
                if(len(title.split('--'))==2):
                    'we dont want to store titleName --valid so her we removeing th --valid that if it exist'
                    splittedTitle=title.split('--')
                
                    if(splittedTitle[1].lower()== 'valid'):
                        data[splittedTitle[0]] = cell.value
                    else:
                        data[title] = cell.value

                
                else:
                    data[title] = cell.value


        datas["usersInfo"].append(data)
        cell_line+=1


    # json model field in django doest accept set so i changed it to list
    datas['useValidation'] = list(datas['useValidation'])
    # print(datas)
    return datas