from Dueapp import models
from rel8.celery import app
from . import  models
from openpyxl import Workbook
from account.models import user as user_related_models
from django.core.files import File



@app.task()
def finicial_report():
    "all this function is gettting memebers data and converting it to excel and dumping it int Financial_and_nonFinancialMembersRecord Model"


    # sheet = workbook.active
    list_of_members_finicial = user_related_models.Memeber.objects.filter(is_financial=True).values_list(
    "user__email","is_exco",'usermemberinfo__name','usermemberinfo__value')
    list_of_members_none_finicial = user_related_models.Memeber.objects.filter(is_financial=False).values_list(
    "user__email","is_exco",'usermemberinfo__name','usermemberinfo__value')

    create_report(list_of_members_finicial,f'Financial Report',for_financial=True)
    create_report(list_of_members_finicial,f'Non-Financial Report',for_financial=False)


def create_report(list_of_members,title,for_financial):
    headers =["Email","Exco Postion",'Label','Value']
    filename =f"{title}.xlsx"
    workbook = Workbook()
    excel_data =[
        headers,
        *list_of_members
    ]

    if excel_data:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=filename)
        for line in excel_data:
            ws.append(line)
        wb.save(filename=filename)
        finicial_report = models.Financial_and_nonFinancialMembersRecord.objects.create(
            name =title,
            for_financial =for_financial
        )

        finicial_report.file=File(open(filename,mode='rb'),name=filename)
        finicial_report.save() 